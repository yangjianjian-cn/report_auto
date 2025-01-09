import logging
import re
from collections import OrderedDict

from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')


# 1-passed 2-failed 3-na 4-nottested

def write_analysis_tocsv(output_file: str, insert_rownum: int, levels: map, result_dicts: dict):
    # level_map初始化
    # levels: {'level1': {1}, 'level2': {1, 3}, 'level3': {1, 3}, 'level4': {1}}
    level_map: map = {}
    # level_map赋值
    for level, status_set in levels.items():
        if 1 in status_set:
            level_map[level] = "passed"
        elif 2 in status_set:
            level_map[level] = "failed"
        elif 3 in status_set:
            level_map[level] = "n/a"
        elif 4 in status_set:
            level_map[level] = "not tested"
    logging.info(f"level_map:{level_map}")

    # excel文件中，一行数据
    for result in result_dicts:
        ordered_result = OrderedDict()
        for key in result:
            if key in ['level2', 'level3', 'level4', 'level5']:
                new_key = f'{key}_result'

                idx = int(extract_numbers(key)) - 1
                map_idx = f'level{idx}'

                if map_idx in level_map:
                    ordered_result[new_key] = level_map[map_idx]
                else:
                    ordered_result[new_key] = "not tested"

            ordered_result[key] = result[key]

        # 替换原来的字典
        result.clear()
        result.update(ordered_result)

    # 按指定行号写入文件内容
    wb = load_workbook(filename=output_file, keep_vba=True)
    ws = wb['IO Checklist']
    # 将数据写入指定行
    start_column = 2  # C列对应数字3
    for col_num, value in enumerate(result.values(), start=start_column):
        ws.cell(row=insert_rownum, column=col_num).value = value  # 注意这里使用了ws而不是sheet
    wb.save(filename=output_file)
    wb.close()
    return output_file


def extract_numbers(s):
    # 使用正则表达式找到所有数字
    numbers = re.findall(r'\d+', s)
    # 将所有找到的数字连接成一个字符串
    return ''.join(numbers)
