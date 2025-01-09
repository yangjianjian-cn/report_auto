__coding__ = "utf-8"

from openpyxl import load_workbook


def find_first_empty_row_after_string(filename, search_string):
    # 加载工作簿
    workbook = load_workbook(filename=filename, read_only=True)
    # 选择工作表
    sheet = workbook.active  # 或者指定 sheet 名称：workbook['SheetName']

    # 初始化行号变量
    start_row = None

    # 遍历所有行
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == search_string:
                start_row = cell.row
                break
        if start_row is not None:
            break

    current_row = start_row
    if start_row is not None:
        # print(f'"{search_string}" found at row {start_row}')

        # 从找到的目标行开始向下检查
        current_row = start_row + 1
        while True:
            # 获取当前行的前10个单元格
            cells_in_row = [sheet.cell(row=current_row, column=col) for col in range(1, 11)]

            # 检查当前行的前10列是否有非空单元格
            has_content = any(cell.value is not None for cell in cells_in_row)

            if not has_content:
                # First completely empty row (in the first 10 columns) after "{search_string}" is row {current_row}'
                break

            # 移动到下一行
            current_row += 1
    else:
        raise Exception(f'"{search_string}" not found in the worksheet')

    return current_row

# 示例调用
# filename = r'C:\Users\Administrator\Downloads\template\IOTest_Man_Tmplt.xlsm'
# search_string = 'analogue input'
# start_row = find_first_empty_row_after_string(filename, search_string)
# print(f'Start row: {start_row}')
