import logging

from openpyxl import load_workbook

from constant.IOTestLevelData import i_a_app1_level1_attribute, i_a_app1_level2_data, i_a_app1_level2_attribute, \
    i_a_app1_level3_data, i_a_app1_level3_attribute, \
    i_a_app1_level4_data, i_a_app1_level4_attribute, i_a_app1_level1_data
from pojo.MSTReqPOJO import ReqPOJO
from tools.common.dat_csv_common import create_file_path

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

"""
output_path: / outputpath / 测试团队 / 测试区域
"""


def analogue_input_report(req_data: ReqPOJO, level1: str, level2: str, level3: str, level4: str) -> str:
    logging.info(f"level1-4: {level1},{level2},{level3},{level4}")
    # / outputpath / 测试团队 / 测试区域
    template_file_name, template_path = create_file_path(req_data.template_name, 'xlsm', req_data.template_path,
                                                         'template')
    logging.info(f"模板目录:{template_path}, 模板文件:{template_file_name}")

    # 输出文件
    output_file_name, output_path = create_file_path(req_data.template_name, 'xlsm', req_data.output_path, 'xlsm')
    logging.info(f"输出目录:{output_path}, 输出文件:{output_file_name}")

    wb = load_workbook(filename=template_path)
    ws = wb['IO Checklist']

    row_index = 9
    columns = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12]  # 13

    # level1
    for i, column in enumerate(columns, start=0):
        attr = getattr(i_a_app1_level1_data, i_a_app1_level1_attribute[i])
        logging.info(f"{column}:{attr}")
        ws.cell(row=row_index, column=column).value = attr
    ws[f'M{row_index}'].value = level1

    # level2
    columns = [14, 15, 16, 17, 18]  # 19
    for i, column in enumerate(columns, start=0):
        attr = getattr(i_a_app1_level2_data, i_a_app1_level2_attribute[i])
        logging.info(f"{column}:{attr}")
        ws.cell(row=row_index, column=column).value = attr
    ws[f'S{row_index}'].value = level2

    # level3
    columns = [20, 21, 22, 23, 24]  # 25
    for i, column in enumerate(columns, start=0):
        attr = getattr(i_a_app1_level3_data, i_a_app1_level3_attribute[i])
        logging.info(f"{column}:{attr}")
        ws.cell(row=row_index, column=column).value = attr
    ws[f'Y{row_index}'].value = level3

    # level4
    columns = [26, 27, 28, 29, 30]  # 31
    for i, column in enumerate(columns, start=0):
        attr = getattr(i_a_app1_level4_data, i_a_app1_level4_attribute[i])
        logging.info(f"{column}:{attr}")
        ws.cell(row=row_index, column=column).value = attr
    ws[f'AE{row_index}'].value = level4

    wb.save(filename=output_path)
    wb.close()
    return output_path
