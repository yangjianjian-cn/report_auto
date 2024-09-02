from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# 加载工作簿
wb = load_workbook(filename='E:\\python_ws\\dmx\\dmx_zcy\\report_auto\\static\\file\\IOTest_Man_Tmplt.xlsm')

# 选择工作表
ws = wb['IO Checklist']


# 设置A1单元格的默认值为Option3
ws['M7'].value = 'passed'

# 保存工作簿
wb.save('E:\\python_ws\\dmx\\dmx_zcy\\report_auto\\static\\file\\IOTest_Man_Tmplt.xlsm')