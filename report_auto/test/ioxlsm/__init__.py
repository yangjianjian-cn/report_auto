from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule

# 加载工作簿
wb = load_workbook(filename='E:\\python_ws\\dmx\\dmx_zcy\\report_auto\\static\\file\\IOTest_Man_Tmplt.xlsm')

# 选择工作表
ws = wb['Sheet1']

# 清除A1单元格的内容
ws['A1'].value = None

# 创建一个新的数据验证对象
dv = DataValidation(type="list", formula1='"Option1,Option2,Option3"')

# 将数据验证应用到特定的单元格
dv.add('A1')

# 设置A1单元格的默认值为Option3
ws['A1'].value = 'Option3'

# 清除原有的数据验证
for dv_existing in ws.data_validations:
    if isinstance(dv_existing, DataValidation) and 'A1' in dv_existing.ranges[0]:
        ws.data_validations.remove(dv_existing)

# 添加新的数据验证
ws.add_data_validation(dv)

# 设置条件格式
yellow_fill = PatternFill(start_color="FFFF00",
                          end_color="FFFF00",
                          fill_type="solid")

red_fill = PatternFill(start_color="FF0000",
                       end_color="FF0000",
                       fill_type="solid")

gray_fill = PatternFill(start_color="D3D3D3",
                        end_color="D3D3D3",
                        fill_type="solid")

# 创建条件格式规则
# 当单元格值等于"Option3"时，背景颜色变为黄色
rule_yellow = CellIsRule(operator='equal', formula=['"Option3"'], stopIfTrue=True, fill=yellow_fill)

# 当单元格值等于"Option1"时，背景颜色变为红色
rule_red = CellIsRule(operator='equal', formula=['"Option1"'], stopIfTrue=True, fill=red_fill)

# 当单元格值等于"Option2"时，背景颜色变为灰色
rule_gray = CellIsRule(operator='equal', formula=['"Option2"'], stopIfTrue=True, fill=gray_fill)

# 将条件格式应用于A1单元格
ws.conditional_formatting.add('A1', rule_yellow)
ws.conditional_formatting.add('A1', rule_red)
ws.conditional_formatting.add('A1', rule_gray)

# 保存工作簿
wb.save('E:\\python_ws\\dmx\\dmx_zcy\\report_auto\\static\\file\\IOTest_Man_Tmplt.xlsm')